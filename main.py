# import getpass
import datetime
import difflib
import os
import re
import sys

import nltk
import openpyxl
import pandas as pd
import pyodbc
import sqlparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

# Colors
RED = "\033[91m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
RESET = "\033[0m"


def ignore(ign_filepath: str, orig_filelist) -> list[str]:
    with open(ign_filepath, "r") as f:
        ignore: list[str]= f.read().splitlines()

    additional_ignore_list = [
        '*.mig*', '*sqlquery*', '*2022*', '*2023*', '*backup*'
    ]

    ignore.extend(additional_ignore_list)
    ignore_patterns_list = [r".*_" + item.upper() + ".*" for item in ignore]


    # Create a regular expression pattern to match ignore patterns
    ignore_pattern = "|".join(ignore_patterns_list)
    ignore_pattern = f"({ignore_pattern})"

    filtered_filelist = [file for file in orig_filelist if not re.match(ignore_pattern, file.upper())]

    return filtered_filelist


def connect_to_server(server, database, username, password):
    try:
        connection_string = f"DRIVER=SQL Server;SERVER={server};DATABASE={database};UID={username};PWD={password}"
        connection = pyodbc.connect(connection_string)
        # cursor = connection.cursor()
    except Exception as e:
        print(f"Error: {str(e)}")
        
    return connection

def list_exception_tables(server, database, username, password) -> list:
    connection = connect_to_server(server, database, username, password)
    cursor = connection.cursor()
    query = """
    SELECT TABLE_SCHEMA, TABLE_NAME
    FROM INFORMATION_SCHEMA.TABLES
    WHERE TABLE_SCHEMA NOT LIKE 'dbo'
    """
    cursor.execute(query)
    except_tables_list: list = []
    except_tables_list = cursor.fetchall()
    
    return except_tables_list

def fetch_schema(server, database, username, password) -> dict:
    connection = connect_to_server(server, database, username, password)
    cursor = connection.cursor()
    schema_info = {}
    
    # AND c.COLUMN_NAME <> 'S# No#'
    # AND t.TABLE_NAME <> 'RCP_BankMaster'
    # AND t.TABLE_NAME <> 'RCP_Excel_Caste'
    # AND t.TABLE_NAME <> 'RCP_Excel_City'
       
    query = """
    SELECT
    LTRIM(RTRIM(t.TABLE_SCHEMA)) AS TABLE_SCHEMA,
    LTRIM(RTRIM(t.TABLE_NAME)) AS TABLE_NAME,
    LTRIM(RTRIM(c.COLUMN_NAME)) AS COLUMN_NAME,
    c.DATA_TYPE,
    c.CHARACTER_MAXIMUM_LENGTH,
    c.NUMERIC_PRECISION,
    c.NUMERIC_SCALE
    FROM INFORMATION_SCHEMA.TABLES AS t
    JOIN INFORMATION_SCHEMA.COLUMNS AS c ON t.TABLE_SCHEMA = c.TABLE_SCHEMA AND
    t.TABLE_NAME = c.TABLE_NAME
    WHERE t.TABLE_TYPE = 'BASE TABLE'
    AND t.TABLE_SCHEMA LIKE 'dbo'
    AND
    (
    t.TABLE_NAME NOT LIKE '%BKUP%'
    AND t.TABLE_NAME NOT LIKE '%BKP%'
    AND t.TABLE_NAME NOT LIKE '%20%'
    AND t.TABLE_NAME NOT LIKE '%SWAPNIL%'
    AND t.TABLE_NAME NOT LIKE '%SQLQUERY%'
    AND t.TABLE_NAME NOT LIKE '%FARHEEN%'
    AND t.TABLE_NAME NOT LIKE '%SHUBHAM%'
    AND t.TABLE_NAME NOT LIKE '%CHHAGAN%'
    AND t.TABLE_NAME NOT LIKE '%TCKT%'
    AND t.TABLE_NAME NOT LIKE '%MIGRATION%'
    AND t.TABLE_NAME NOT LIKE '%MIGR%'
    AND t.TABLE_NAME NOT LIKE '%TID%'
    AND t.TABLE_NAME NOT LIKE '%tblPivoPOAttainmet%'
    AND t.TABLE_NAME NOT LIKE '%BK%'
    AND t.TABLE_NAME NOT LIKE '%BACKUP%'
    AND t.TABLE_NAME NOT LIKE '%TKT%'
    AND t.TABLE_NAME NOT LIKE '%TICKET_ID%'
    AND t.TABLE_NAME NOT LIKE '%TICKET%'
    AND t.TABLE_NAME NOT LIKE '%MIG%'
    AND t.TABLE_NAME NOT LIKE '%_BKUP%'
    AND t.TABLE_NAME NOT LIKE '%_BKP%'
    AND t.TABLE_NAME NOT LIKE '%20%'
    AND t.TABLE_NAME NOT LIKE '%SWAPNIL%'
    AND t.TABLE_NAME NOT LIKE '%SQLQUERY%'
    AND t.TABLE_NAME NOT LIKE '%FARHEEN%'
    AND t.TABLE_NAME NOT LIKE '%SHUBHAM%'
    AND t.TABLE_NAME NOT LIKE '%CHHAGAN%'
    AND t.TABLE_NAME NOT LIKE '%_TCKT%'
    AND t.TABLE_NAME NOT LIKE '%1_9%'
    AND t.TABLE_NAME NOT LIKE '%$%'
    AND t.TABLE_NAME NOT LIKE '%_TID%'
    AND t.TABLE_NAME NOT LIKE '%tblPivoPOAttainmet%'
    AND t.TABLE_NAME NOT LIKE '%_BK%'
    AND t.TABLE_NAME NOT LIKE '%_BAK%'
    AND t.TABLE_NAME NOT LIKE '%BACKUP%'
    AND t.TABLE_NAME NOT LIKE '%TKT%'
    AND t.TABLE_NAME NOT LIKE '%22%'
    AND t.TABLE_NAME NOT LIKE '%TICKET_ID%'
    AND t.TABLE_NAME NOT LIKE '%TICKET%'
    AND t.TABLE_NAME NOT LIKE '%lms%'
    AND t.TABLE_NAME NOT LIKE '%123%'
    AND t.TABLE_NAME NOT LIKE '%aarv%'
    AND t.TABLE_NAME NOT LIKE '%BS23%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%amtt%'
    AND t.TABLE_NAME NOT LIKE '%5186%'
    AND t.TABLE_NAME NOT LIKE '%ms21%'
    AND t.TABLE_NAME NOT LIKE '%Com%'
    AND t.TABLE_NAME NOT LIKE '%DA23%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%c001%'
    AND t.TABLE_NAME NOT LIKE '%MCA%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%5186%'
    AND t.TABLE_NAME NOT LIKE '%BE%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%5186%'
    AND t.TABLE_NAME NOT LIKE '%ms21%'
    AND t.TABLE_NAME NOT LIKE '%DA23%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%DA23%'
    AND t.TABLE_NAME NOT LIKE '%RC21%'
    AND t.TABLE_NAME NOT LIKE '%_BKUP%'
    AND t.TABLE_NAME NOT LIKE '%_BKP%'
    AND t.TABLE_NAME NOT LIKE '%20%'
    AND t.TABLE_NAME NOT LIKE '%SWAPNIL%'
    AND t.TABLE_NAME NOT LIKE '%SQLQUERY%'
    AND t.TABLE_NAME NOT LIKE '%FARHEEN%'
    AND t.TABLE_NAME NOT LIKE '%SHUBHAM%'
    AND t.TABLE_NAME NOT LIKE '%CHHAGAN%'
    AND t.TABLE_NAME NOT LIKE '%_TCKT%'
    AND t.TABLE_NAME NOT LIKE '%1_9%'
    AND t.TABLE_NAME NOT LIKE '%$%'
    AND t.TABLE_NAME NOT LIKE '%_TID%'
    AND t.TABLE_NAME NOT LIKE '%tblPivoPOAttainmet%'
    AND t.TABLE_NAME NOT LIKE '%_BK%'
    AND t.TABLE_NAME NOT LIKE '%_BAK%'
    AND t.TABLE_NAME NOT LIKE '%BACKUP%'
    AND t.TABLE_NAME NOT LIKE '%TKT%'
    AND t.TABLE_NAME NOT LIKE '%22%'
    AND t.TABLE_NAME NOT LIKE '%TICKET_ID%'
    AND t.TABLE_NAME NOT LIKE '%TICKET%'
    AND t.TABLE_NAME NOT LIKE '%lms%'
    AND t.TABLE_NAME NOT LIKE '%123%'
    AND t.TABLE_NAME NOT LIKE '%aarv%'
    AND t.TABLE_NAME NOT LIKE '%BS23%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%amtt%'
    AND t.TABLE_NAME NOT LIKE '%5186%'
    AND t.TABLE_NAME NOT LIKE '%ms21%'
    AND t.TABLE_NAME NOT LIKE '%Com%'
    AND t.TABLE_NAME NOT LIKE '%DA23%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%c001%'
    AND t.TABLE_NAME NOT LIKE '%MCA%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%5186%'
    AND t.TABLE_NAME NOT LIKE '%BE%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%5186%'
    AND t.TABLE_NAME NOT LIKE '%ms21%'
    AND t.TABLE_NAME NOT LIKE '%DA23%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%DA23%'
    AND t.TABLE_NAME NOT LIKE '%01%'
    AND t.TABLE_NAME NOT LIKE '%RC21%'
    AND t.TABLE_NAME NOT LIKE '%gaag%'
    AND t.TABLE_NAME NOT LIKE '%RM1%'
    AND t.TABLE_NAME NOT LIKE '%ms55%'
    AND t.TABLE_NAME NOT LIKE '%MC21%'
    AND t.TABLE_NAME NOT LIKE '%C23%'
    AND t.TABLE_NAME NOT LIKE '%BSACIST%'
    AND t.TABLE_NAME NOT LIKE '%BSAC%'
    AND t.TABLE_NAME NOT LIKE '%M01%'
    AND t.TABLE_NAME NOT LIKE '%BSACIST%'
    AND t.TABLE_NAME NOT LIKE '%TBAK%'
    AND t.TABLE_NAME NOT LIKE '%BS23%'
    AND t.TABLE_NAME NOT LIKE '%BE%'
    AND t.TABLE_NAME NOT LIKE '%MBA%'
    AND t.TABLE_NAME NOT LIKE '%MCA%'
    AND t.TABLE_NAME NOT LIKE '%ME%'
    AND t.TABLE_NAME NOT LIKE '%TEST%'
    AND t.TABLE_NAME NOT LIKE '%DEV%'
    AND t.TABLE_NAME NOT LIKE '%BEPT%'
    AND t.TABLE_NAME NOT LIKE '%HITS%'
    AND t.TABLE_NAME NOT LIKE '%BSACIST%'
    AND t.TABLE_NAME NOT LIKE '%ms21%'
    AND t.TABLE_NAME NOT LIKE '%test%'
    AND t.TABLE_NAME NOT LIKE '%c001%'
    AND t.TABLE_NAME NOT LIKE '%m01%'
    AND t.TABLE_NAME NOT LIKE '%CR23%'
    AND t.TABLE_NAME NOT LIKE '%C23%'
    AND t.TABLE_NAME NOT LIKE '%123%'
    AND t.TABLE_NAME NOT LIKE '%lms%'
    AND t.TABLE_NAME NOT LIKE '%AARV%'
    AND t.TABLE_NAME NOT LIKE '%BE%'
    AND t.TABLE_NAME NOT LIKE '%MBA%'
    AND t.TABLE_NAME NOT LIKE '%MCA%'
    AND t.TABLE_NAME NOT LIKE '%ME%'
    AND t.TABLE_NAME NOT LIKE '%TEST%'
    AND t.TABLE_NAME NOT LIKE '%DEV%'
    AND t.TABLE_NAME NOT LIKE '%BEPT%'
    AND t.TABLE_NAME NOT LIKE '%BE%'
    AND t.TABLE_NAME NOT LIKE '%MBA%'
    AND t.TABLE_NAME NOT LIKE '%MCA%'
    AND t.TABLE_NAME NOT LIKE '%ME%'
    AND t.TABLE_NAME NOT LIKE '%TEST%'
    AND t.TABLE_NAME NOT LIKE '%DEV%'
    AND t.TABLE_NAME NOT LIKE '%BEPT%'
    AND t.TABLE_NAME NOT LIKE '%DA23%'
    AND t.TABLE_NAME NOT LIKE '%MAHR%'
    AND t.TABLE_NAME NOT LIKE '%sims%'
    AND t.TABLE_NAME NOT LIKE '%sim%'
    AND t.TABLE_NAME NOT LIKE '%S01%'
    AND t.TABLE_NAME NOT LIKE '%S02%'
    
    
    AND t.TABLE_NAME NOT LIKE '%ACC_01%'
    AND t.TABLE_NAME NOT LIKE '%ACC_BE%'
    AND t.TABLE_NAME NOT LIKE '%MS21%'
    AND t.TABLE_NAME NOT LIKE '%C001%'
    AND t.TABLE_NAME NOT LIKE '%ACC_TEST%'
    AND t.TABLE_NAME NOT LIKE '%TCKT%'
    AND t.TABLE_NAME NOT LIKE '%48069%'
    AND t.TABLE_NAME NOT LIKE '%acc_%'
    AND t.TABLE_NAME NOT LIKE '%MIGR%'
    AND t.TABLE_NAME NOT LIKE '%MIGRATION%'
    AND t.TABLE_NAME NOT LIKE '%PAY_ITMTEMP%'
    AND t.TABLE_NAME NOT LIKE '%PAY_PAYTEMP%'
    AND t.TABLE_NAME NOT LIKE '%PAY_ITET%'
    AND t.TABLE_NAME NOT LIKE '%TEMP%'
    AND t.TABLE_NAME NOT LIKE '%TID%'
    AND t.TABLE_NAME NOT LIKE '%_tkt%'
    AND t.TABLE_NAME NOT LIKE '%ckt_%'
    AND t.TABLE_NAME NOT LIKE '%ckt_%'
    AND t.TABLE_NAME NOT LIKE '%user_acc_back%'
    AND t.TABLE_NAME NOT LIKE '%user_acc_superadmin%'
    AND t.TABLE_NAME NOT LIKE 'FINAL_PAYSLIP_DATA'
    
    )
    ORDER BY t.TABLE_NAME
    """
    # # ORDER BY t.TABLE_SCHEMA, t.TABLE_NAME, c.ORDINAL_POSITION
    # try removing the above line
    print("\nExecuting the schema query. Please wait...\n")
    cursor.execute(query)
    print("Schema query execution" + GREEN + " completed" + RESET + ".\n")
    # rows = cursor.fetchall()
    # rows = cursor.fetchmany(1000)
    batchsize: int = 1000
    batch = cursor.fetchmany(batchsize)
    
    while batch:
        for row in batch:
            table_schema = row.TABLE_SCHEMA
            table_name = row.TABLE_NAME
            column_name = row.COLUMN_NAME
            data_type = row.DATA_TYPE
            max_length = row.CHARACTER_MAXIMUM_LENGTH
            numeric_precision = row.NUMERIC_PRECISION
            numeric_scale = row.NUMERIC_SCALE


            schema_info.setdefault(table_schema, {}).setdefault(table_name, []).append(
                {
                    "column_name": column_name,
                    "data_type": data_type,
                    "max_length": max_length,
                    "numeric_precision": numeric_precision,
                    "numeric_scale": numeric_scale,
                }
            )
        batch = cursor.fetchmany(batchsize)
    cursor.close()
    connection.close()
    return schema_info



def generate_excel_report(
    workbook,
    source_schema,
    target_schema,
    target_db_name
    ):
    
    comparison_results = []
    try:
        comparison_results = perform_schema_comparison(source_schema, target_schema)
        # print(comparison_results)
    except Exception as e:
        print(f"ERROR: could not compare schemas successfully: {str(e)}")
    
    sheet = workbook.create_sheet(target_db_name)
    # workbook.active = workbook.sheetnames.index(target_db_name)
    
    sheet["A1"] = "Schema"
    sheet["B1"] = "Table Name"
    sheet["C1"] = "Column Name"
    sheet["D1"] = "Type of Error"
    sheet["E1"] = "Source Specification"
    sheet["F1"] = "Target Specification"
    
    sheet["G1"] = "Source Column Name"
    sheet["H1"] = "Source Data Type"
    sheet["I1"] = "Source Max Length"
    sheet["J1"] = "Source Numeric Precision"
    sheet["K1"] = "Source Numeric Scale"
    
    sheet["L1"] = "Target Column Name"
    sheet["M1"] = "Target Data Type"
    sheet["N1"] = "Target Max Length"
    sheet["O1"] = "Target Numeric Precision"
    sheet["P1"] = "Target Numeric Scale"


    row = 2
    for result in comparison_results:
        sheet[f"A{row}"] = result["schema"]
        sheet[f"B{row}"] = result["table_name"]
        sheet[f"C{row}"] = result["column_name"]
        sheet[f"D{row}"] = result["type_of_error"]
        # TODO: change names
        sheet[f"E{row}"] = result["source_specification"]
        sheet[f"F{row}"] = result["target_specification"]
        
        # Changes
        sheet[f"G{row}"] = result['source_column_name']
        sheet[f"H{row}"] = result['source_data_type']
        sheet[f"I{row}"] = result['source_max_length']
        sheet[f"J{row}"] = result['source_numeric_precision']
        sheet[f"K{row}"] = result['source_numeric_scale']

        sheet[f"L{row}"] = result['target_column_name']
        sheet[f"M{row}"] = result['target_data_type']
        sheet[f"N{row}"] = result['target_max_length']
        sheet[f"O{row}"] = result['target_numeric_precision']
        sheet[f"P{row}"] = result['target_numeric_scale']

        if result["type_of_error"] == "Missing Column":
            sheet[f"D{row}"].fill = PatternFill(
                start_color="F9B9B7", end_color="F9B9B7", fill_type="solid"
            )
        elif result["type_of_error"] == "Different Specification":
            sheet[f"D{row}"].fill = PatternFill(
                start_color="D1D5DE", end_color="D1D5DE", fill_type="solid"
            )
        elif result["type_of_error"] == "Missing Table":
            sheet[f"D{row}"].fill = PatternFill(
                start_color="c4d79b", end_color="c4d79b", fill_type="solid"
            )
        row = row + 1


def perform_schema_comparison(source_schema, target_schema) -> list[dict]:
    comparison_results: list[dict] = []
    # print(f"Source Schema: {source_schema}")
    for schema in source_schema:
        # print(f"Schema in source schema: {schema}")
        for table_name in source_schema[schema]:
            print(f"Table Name: {table_name}")
            try:
                source_columns = source_schema[schema][table_name]
            except Exception as e:
                print(f"Failed while setting source solumns\n {table_name}")
                print(f"Exception: {str(e)}")
            # print(f"Source Columns: {source_columns}")
            try:
                target_columns = target_schema[schema].get(table_name, [])
            except Exception as e:
                print(f"Failed while setting target columns {table_name}")
                print(f"Exception: {str(e)}")
            source_columns = source_schema[schema][table_name]
            target_columns = target_schema[schema].get(table_name, [])
            # print(f"Target Columns: {target_columns}")
            # print(f"Comparing {source_columns} with {target_columns}")

            if target_columns is None:
                print(f"{table_name} Table Not found in target database")
                comparison_result = {
                    "schema": schema,
                    "table_name": table_name,
                    "column_name": f"Table {table_name} Not Found",
                    "type_of_error": "Missing Table",
                    "source_specification": "",
                    "target_specification": "",
                    # Changes
                    "source_column_name": "",
                    "source_data_type": "",
                    "source_max_length": "",
                    "source_numeric_precision": "",
                    "source_numeric_scale": "",
                    # Target will be empty
                    "target_column_name": "",
                    "target_data_type": "",
                    "target_max_length": "",
                    "target_numeric_precision": "",
                    "target_numeric_scale": "",
                }
                # print(f"{comparison_result}\n\n\n")
                try:
                    comparison_results.append(comparison_result)
                except Exception as e:
                    print(f"Failed inside table col none for {target_columns}")
                    print(f"Exception: {str(e)}\n\n\n")
                # print(f"{comparison_results}\n\n\n")
            else:
                for col_info_source in source_columns:
                    # print(f"col_info_source: {col_info_source}")
                    col_name_source = col_info_source["column_name"]
                    # print(f"col_name_source: {col_name_source}")
                    col_info_target = next(
                        (col for col in target_columns if col["column_name"] == col_name_source),
                        None,
                    )
                    # print(f"col_info_target: {col_info_target}")
                    if col_info_target is None:
                        # Missing Column
                        comparison_result = {
                            "schema": schema,
                            "table_name": table_name,
                            "column_name": col_name_source,
                            "type_of_error": "Missing Column",
                            "source_specification": str(col_info_source),
                            "target_specification": "",
                            # CHANGES
                            "source_column_name": col_info_source['column_name'],
                            "source_data_type": col_info_source['data_type'],
                            "source_max_length": col_info_source['max_length'],
                            "source_numeric_precision": col_info_source['numeric_precision'],
                            "source_numeric_scale": col_info_source['numeric_scale'],
                            # Target will be empty ideally
                            "target_column_name": "",
                            "target_data_type": "",
                            "target_max_length": "",
                            "target_numeric_precision": "",
                            "target_numeric_scale": "",
                        }
                        # print(f"{comparison_result}\n\n\n")
                        try:
                            comparison_results.append(comparison_result)
                        except Exception as e:
                            print(f"Failed inside col info target none for {col_info_source}")
                            print(f"Exception: {str(e)}\n\n\n")
                        # print(f"{comparison_results}\n\n\n")

                    # Different Specification
                    elif col_info_source != col_info_target:
                        comparison_result = {
                            "schema": schema,
                            "table_name": table_name,
                            "column_name": col_name_source,
                            "type_of_error": "Different Specification",
                            "source_specification": str(col_info_source),
                            "target_specification": str(col_info_target),
                            # Changes
                            "source_column_name": col_info_source['column_name'],
                            "source_data_type": col_info_source['data_type'],
                            "source_max_length": col_info_source['max_length'],
                            "source_numeric_precision": col_info_source['numeric_precision'],
                            "source_numeric_scale": col_info_source['numeric_scale'],
                            # Target
                            "target_column_name": col_info_target['column_name'],
                            "target_data_type": col_info_target['data_type'],
                            "target_max_length": col_info_target['max_length'],
                            "target_numeric_precision": col_info_target['numeric_precision'],
                            "target_numeric_scale": col_info_target['numeric_scale'],
                        }
                        # print(f"{comparison_result}\n\n\n")
                        try:
                            comparison_results.append(comparison_result)
                        except Exception as e:
                            print(f"Failed inside col != col for {col_info_source}")
                            print(f"Exception: {str(e)}")
                        # print(f"{comparison_results}\n\n\n")
                    elif col_info_source == col_info_target:
                        # print("cols equal continuing\n\n\n")
                        continue
                    else:
                        print("What case even is this??\n\n\n")
                        continue
    # print(f"{comparison_results}\n\n\n")
    return comparison_results


def app1() -> None:
    try:
        print("Enter details of" + GREEN + " SOURCE " + RESET + "database: ")
        source_info = {
            "server": 1234,
            "database": "dbdb",
            "username": "uid",
            "password": "pwd",
        }

        source_info["server"] = input(
            "Enter" + GREEN + " Server " + RESET + "Address:\t"
        )
        source_info["database"] = input(
            "Enter" + GREEN + " Database " + RESET + "Name:\t"
        )
        source_info["username"] = input("Enter Server" + GREEN + " Username:\t" + RESET)
        source_info["password"] = input("Enter Server" + GREEN + " Password:\t" + RESET)

        source_server = source_info["server"]
        source_database = source_info["database"]
        source_username = source_info["username"]
        source_password = source_info["password"]

        if (
            not source_server
            or not source_database
            or not source_username
            or not source_password
        ):
            raise ValueError(
                "Source database details are"
                + RED
                + " incomplete"
                + RESET
                + ". Please provide all required information."
            )

        print("Fetching Source Information... \n")


        try:
            source_schema = fetch_schema(source_server, source_database, source_username, source_password)
            print(type(source_schema))
            
            debug_source_schema_file = "ss.txt"
            with open(debug_source_schema_file, 'w') as dts:
                dts.write(str(source_schema))
            # print(f"Source Schema:\n{str(source_schema)}")
            
        except Exception:
            raise ValueError(
                "Couldn't fetch source schema. Perhaps the"
                + RED
                + " login details "
                + RESET
                + "or"
                + RED
                + " database details "
                + RESET
                + "you entered are incorrect?\n"
            )

        workbook = openpyxl.Workbook()

        try:
            number_of_target_db = input(
                "How many databases do you want to compare against source? \t"
            )
            number_of_target_db = int(number_of_target_db)
        except Exception:
            print(
                RED
                + "Error: "
                + RESET
                + "No input or incorrect form of input was provided."
            )
            return

        nested_target_dbs = {}
        number_of_target_db_copy = number_of_target_db

        if number_of_target_db <= 0:
            print("No databases to compare." + RED + " Exiting." + RESET)
            return

        for target_db_number in range(1, number_of_target_db + 1):
            nested_target_dbs[target_db_number] = {}
            nested_target_dbs[target_db_number]["server"] = input(
                "\nEnter"
                + GREEN
                + " Server Address "
                + RESET
                + f"for Target DB number {target_db_number}: "
            )
            nested_target_dbs[target_db_number]["database"] = input(
                "Enter"
                + GREEN
                + " Database Name "
                + RESET
                + f"for Target DB number {target_db_number}: "
            )
            nested_target_dbs[target_db_number]["username"] = input(
                "Enter"
                + GREEN
                + " User Name "
                + RESET
                + f"for Target DB number {target_db_number}: "
            )
            nested_target_dbs[target_db_number]["password"] = input(
                "Enter"
                + GREEN
                + " Password "
                + RESET
                + f"for Target DB number {target_db_number}: "
            )

            target_server = nested_target_dbs[target_db_number]["server"]
            target_database = nested_target_dbs[target_db_number]["database"]
            target_username = nested_target_dbs[target_db_number]["username"]
            target_password = nested_target_dbs[target_db_number]["password"]

            if (
                not target_server
                or not target_database
                or not target_username
                or not target_password
            ):
                print(
                    f"Target database number {target_db_number}'s details are"
                    + RED
                    + " incomplete"
                    + RESET
                    + ". Skipping comparison for this target."
                )
            else:
                try:
                    target_schema = fetch_schema(target_server, target_database, target_username, target_password)
                    print(type(target_schema))
                    debug_target_schema_file = "ts.txt"
                    with open(debug_target_schema_file, 'w') as dts:
                        dts.write(str(target_schema))
                    try:
                        generate_excel_report(
                            workbook,
                            source_schema,
                            target_schema,
                            nested_target_dbs[target_db_number]["database"]
                        )
                    except Exception as e:
                        print(f"Error generating excel report but target database connection was successful.: {str(e)}")

                except Exception as e:
                    print(
                        "\nError fetching schema for target database number: "
                        + f"{target_db_number}: {str(e)}"
                    )

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
        excel_file_name = f"Schema_Comparison_Report_{timestamp}.xlsx"

        workbook.remove(workbook.active)
        workbook.save(excel_file_name)
        print(
            "\nExcel file"
            + GREEN
            + " successfully "
            + RESET
            + f"created at {os.path.abspath(excel_file_name)}\n\n"
        )

        summary = {}

        excel_file = openpyxl.load_workbook(excel_file_name)

        for sheet_name in excel_file.sheetnames:
            sheet = excel_file[sheet_name]

            missing_columns = 0
            different_specifications = 0
            total_differences = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                data_type = row[3]

                if data_type == "Missing Column":
                    missing_columns = missing_columns + 1
                elif data_type == "Different Specification":
                    different_specifications = different_specifications + 1

            total_differences = missing_columns + different_specifications

            summary[sheet_name] = {
                "Missing Columns": missing_columns,
                "Different Specifications": different_specifications,
                "Total Differences": total_differences,
            }

        excel_file.close()

        print(YELLOW + "Summary: \n" + RESET)

        for target_db, target_summary in summary.items():
            print(YELLOW + "\nTarget Database: " + RESET + f"{target_db}")
            print(f"Missing Columns: {target_summary['Missing Columns']}")
            print(
                f"Different Specifications: {target_summary['Different Specifications']}"
            )
            print(f"Total Differences: {target_summary['Total Differences']}\n")

        print(
            f"\nNumber of databases compared against source: {number_of_target_db_copy}"
        )
        
        try:
            ex_tables = list_exception_tables(source_server, source_database, source_username, source_password)
            print(RED
                  + "\n\n"
                  + "EXCEPTION CAUSING TABLES:"
                  + RESET
                  + f"{ex_tables}\n"
                  + "THESE HAVE BEEN IGNORED BUT "
                  + f"PLEASE RECTIFY THEM IN DATABASE {source_database}\n"
                  + "NO EXCEPTIONS HAVE BEEN CHECKED FOR IN TARGET DATABASE\n\n"
                  )
        except Exception as e:
            print(f"Exception: {str(e)}")
        
        os.system(f'explorer /select,"{os.path.abspath(excel_file_name)}"')
        input("\n\nPress Enter to exit...")
    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Error: {str(e)}")


# END OF APP 1 #########################################################################


def fetch_stored_procedures(server, database, username, password) -> list[str]:
    stored_procedures = []

    try:
        # uses the connect_to_server() function
        connection = connect_to_server(server, database, username, password)
        cursor = connection.cursor()

        # Fetch stored procedures
        query = """
        SELECT name
        FROM sys.procedures
        """

        cursor.execute(query)
        rows = cursor.fetchall()  # to get all stored procedures
        # rows = cursor.fetchmany(1000)

        for row in rows:
            stored_procedures.append(row[0])

        cursor.close()
        connection.close()

    except pyodbc.Error as e:
        print(f"Error: {str(e)}")

    return stored_procedures


# Duplicate strip sql comment func
# def strip_sql_comments(sql_file_contents) -> str:
#     try:
#         stripped_sql_file = re.sub(r"--.*", "", sql_file_contents)
#         stripped_sql_file = re.sub(r"/\*.*?\*/", "", stripped_sql_file, flags=re.DOTALL)
#         stripped_sql_file = "\n".join(
#             " ".join(part.strip() for part in line.split() if part.strip())
#             for line in stripped_sql_file.splitlines()
#             if line.strip()
#         )
#     except Exception as e:
#         print(f"Error while stripping comments: {str(e)}")
#         stripped_sql_file = ""
#     return stripped_sql_file

"""
This function is supposed to ignore lines before the word "Create"
This is useful but has a few caveats such as what if there is the word create
in the comments before the actual statement that creates the procedure
also this does not check the differences on the line itself
Ideally should check after the second square bracket
Example: CREATE      PROCEDURE [PR_ACD_SECTION_MASTER_SHOW_MULTI_MASTERS] abcd
"""


def strip_comments_after_create(sql_file_contents):
    try:
        # Split the SQL content into lines
        lines = sql_file_contents.splitlines()
        create_found = False
        stripped_lines = []

        for line in lines:
            # Check if the line contains "CREATE" (case-insensitive)
            if not create_found and re.search(r"\bCREATE\b", line, re.IGNORECASE):
                create_found = True
                continue

            if create_found:
                line = re.sub(r"--.*", "", line)
                line = re.sub(r"/\*.*?\*/", "", line, flags=re.DOTALL)
                line = line.strip()
                if line:
                    stripped_lines.append(line)

        stripped_sql_file = "\n".join(stripped_lines)

    except Exception as e:
        print(f"Error while stripping comments: {str(e)}")
        stripped_sql_file = ""

    return stripped_sql_file


# The content is not tokenized while checking difference
def difference(source_sql_path, test_sql_path):
    try:
        try:
            with open(source_sql_path, "r", encoding="utf-8") as file:
                source_contents = file.read().upper().strip()
        except UnicodeError:
            with open(source_sql_path, "r", encoding="utf-16") as file:
                source_contents = file.read().upper()

        try:
            with open(test_sql_path, "r", encoding="utf-8") as file:
                test_contents = file.read().upper().strip()
        except UnicodeError:
            with open(test_sql_path, "r", encoding="utf-16") as file:
                test_contents = file.read().upper()

        stripped_sql_file_source = strip_comments_after_create(source_contents)
        stripped_sql_file_test = strip_comments_after_create(test_contents)

        if stripped_sql_file_source == stripped_sql_file_test:
            return True
        else:
            return False

    except Exception as e:
        print(f"Error while comparing SQL Files: {str(e)}")

def app2_4() -> None:
    try:
        print("Enter details of your" + GREEN + " Source Database " + RESET + ":\n")
        source_db_dir = input(
            "Enter the Source Database" + GREEN + " directory location" + RESET + ":\t"
        )
        print("\n")

        if not source_db_dir.strip():
            print(
                RED
                + "Error: "
                + RESET
                + "No source database provided. Exiting program...\n\n"
            )
            sys.exit(1)

        if not os.path.isdir(source_db_dir):
            print(
                RED
                + "Error: "
                + RESET
                + f"Directory '{source_db_dir}' is invalid. Exiting program..."
            )
            sys.exit(1)

        num_target_dbs = input(
            "Enter"
            + GREEN
            + " number of Target Databases "
            + RESET
            + "you want to compare:\t"
        )

        if not num_target_dbs:
            print(
                RED
                + "\n\nError: "
                + RESET
                + "No input provided. Exiting the program...\n\n"
            )
            sys.exit(1)

        try:
            num_target_dbs = int(num_target_dbs)
        except ValueError:
            print(
                RED
                + "\n\nError: "
                + RESET
                + "Invalid input. Please enter a valid positive integer.\n\n"
            )
            sys.exit(1)

        target_db_dirs = []

        for i in range(num_target_dbs):
            while True:
                target_db_dir = input(
                    "\nEnter "
                    + GREEN
                    + "target database directory location "
                    + RESET
                    + f"for target database number {i+1}:\t"
                )

                if not target_db_dir.strip():
                    print(
                        RED
                        + "Error: "
                        + RESET
                        + "No input provided. Please enter a directory location.\n"
                    )
                elif not os.path.exists(target_db_dir):
                    print(
                        RED
                        + "Error: "
                        + RESET
                        + f"The directory '{target_db_dir}' does not exist" + 
                            "or is invalid. Please enter a valid directory location.\n"
                    )
                else:
                    target_db_dirs.append(target_db_dir)
                    break  # Valid input, exit the loop

        sp_data = []

        # Get the list of sql file in source database
        source_sql_file_list = os.listdir(source_db_dir)
        total_files = len(source_sql_file_list)

        ignore_file_path = input("Drag and drop the ignore file:\t")
        filtered_filelist: list[str] = ignore(ignore_file_path, source_sql_file_list)
        
        files_aft_excl = len(filtered_filelist)
        excl_files = total_files - files_aft_excl


        summary = {}

        for target_db_dir in target_db_dirs:
            summary[target_db_dir] = {
                "Absent Entries": 0,
                "Present & Unequal Entries": 0,
                "Present & Equal Entries": 0,
            }

        for sql_file in tqdm(filtered_filelist):
            sp_name = sql_file[:-4]  # Remove the .sql extension
            sp_info = {"SQL File": sp_name}
            source_sql_path = os.path.join(source_db_dir, sql_file)

            for target_db_dir in target_db_dirs:
                target_sql_path = os.path.join(target_db_dir, sql_file)
                if os.path.exists(target_sql_path):
                    if difference(source_sql_path, target_sql_path):
                        sp_info[os.path.basename(target_db_dir)] = "PRESENT & EQUAL"
                        summary[target_db_dir]["Present & Equal Entries"] += 1
                    else:
                        sp_info[os.path.basename(target_db_dir)] = "PRESENT & UNEQUAL"
                        summary[target_db_dir]["Present & Unequal Entries"] += 1
                else:
                    sp_info[os.path.basename(target_db_dir)] = "ABSENT"
                    summary[target_db_dir]["Absent Entries"] += 1

            sp_data.append(sp_info)

        df = pd.DataFrame(sp_data)

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
        output_excel_path = f"SP_Comparison_Report_{timestamp}.xlsx"
        df.to_excel(output_excel_path, index=False)

        wb = load_workbook(output_excel_path)
        ws = wb.active

        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
        ):
            for cell in row:
                if cell.value == "PRESENT & UNEQUAL":
                    cell.fill = PatternFill(
                        start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"
                    )
                elif cell.value == "ABSENT":
                    cell.fill = PatternFill(
                        start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
                    )

        wb.save(output_excel_path)

        print(YELLOW + "\n\nSummary:\n\n" + RESET)
        for target_db_dir, summary_data in summary.items():
            print(
                YELLOW
                + "\nTarget Database: "
                + RESET
                + GREEN
                + f"{os.path.basename(target_db_dir)}"
                + RESET
            )
            print(f"Absent Entries: {summary_data['Absent Entries']}")
            print(
                "Present & Unequal Entries:"
                + f"{summary_data['Present & Unequal Entries']}"
            )
            print("Present & Equal Entries:"
                    + f" {summary_data['Present & Equal Entries']}")
            print(
                "Total Entries Scanned Against Source:"
                + f" {summary_data['Absent Entries']+summary_data['Present & Unequal Entries']+summary_data['Present & Equal Entries']}"
            )
            print("\n")

        excel_absolute_path = os.path.abspath(output_excel_path)
        print(
            GREEN
            + "\nSuccess: "
            + RESET
            + f"Excel Report has been generated at {excel_absolute_path}\n"
        )
        os.system(f'explorer /select, "{os.path.abspath(output_excel_path)}"')

        print("\n\nSummary or Exclusion:\n")
        print(f"Total Files: {total_files}")
        print(f"Included Files: {files_aft_excl}")
        print(f"Excluded Files: {excl_files}")

        input("Press Enter to exit...")

    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")


# END OF APP 2 #########################################################################


def strip_sql_comments(sql_file_contents) -> str:
    # Remove single-line comments (--)
    stripped_sql_file = re.sub(r"--.*", "", sql_file_contents)
    # Remove multi-line comments (/* */)
    stripped_sql_file_1 = re.sub(r"/\*.*?\*/", "", stripped_sql_file, flags=re.DOTALL)
    # Normalize whitespace and line endings
    stripped_sql_file_2 = " ".join(
        line.strip() for line in stripped_sql_file_1.splitlines()
    )
    return stripped_sql_file_2


def normalize_sql(file_contents):
    tokens = nltk.word_tokenize(file_contents)
    file_contents_stringed = " ".join(tokens)
    return file_contents_stringed


def generate_html_diff(file1_contents, file2_contents, folder1_path, folder2_path):
    folder1_name = os.path.basename(folder1_path)
    folder2_name = os.path.basename(folder2_path)

    file1_formatted = normalize_sql(file1_contents)
    file2_formatted = normalize_sql(file2_contents)
    # Remove comments and extra whitespace
    file1_formatted = strip_sql_comments(file1_contents).upper()
    file2_formatted = strip_sql_comments(file2_contents).upper()

    file1_formatted = sqlparse.format(
        file1_formatted, reindent=True, keyword_case="upper", strip_comments=True
    )
    file2_formatted = sqlparse.format(
        file2_formatted, reindent=True, keyword_case="upper", strip_comments=True
    )

    html_diff = difflib.HtmlDiff(tabsize=4, wrapcolumn=72).make_file(
        file1_formatted.splitlines(),
        file2_formatted.splitlines(),
        context=True,
        numlines=5,
        fromdesc=folder1_name,
        todesc=folder2_name,
    )
    return html_diff


def app3():
    comparison_results = []

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
    excel_output_file = f"SP_Comparison_Repost__{timestamp}.xlsx"

    output_dir = f"Diff_Files_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)

    # Source Database
    folder1_path = input(
        "Enter location of the" + GREEN + " Source " + RESET + "database directory: \t"
    )
    if not folder1_path.strip():
        print(
            RED
            + "\nError: "
            + RESET
            + "No source database provided. Exiting program...\n"
        )
        sys.exit(1)

    if not os.path.isdir(folder1_path):
        print(
            RED
            + "\nError: "
            + RESET
            + f"The directory '{folder1_path}' is invalid. Exiting program...\n"
        )
        sys.exit(1)

    # Database to compare with source
    folder2_path = input(
        "Enter location of the" + GREEN + " Target " + RESET + "database directory: \t"
    )
    if not folder2_path.strip():
        print(
            RED
            + "\nError: "
            + RESET
            + "No source database provided. Exiting program...\n"
        )
        sys.exit(1)

    if not os.path.isdir(folder2_path):
        print(
            RED
            + "Error: "
            + RESET
            + f"The directory '{folder2_path}' is invalid. Exiting program...\n"
        )
        sys.exit(1)

    different_color = PatternFill(
        start_color="8ab4ff", end_color="8ab4ff", fill_type="solid"
    )
    missing_color = PatternFill(
        start_color="ffd6ca", end_color="ffd6ca", fill_type="solid"
    )

    # nltk.download("words", quiet=True)
    # nltk.download("punkt", quiet=True)
    # nltk.download("words", quiet=True)
    for sql_file in tqdm(os.listdir(folder1_path)):  # Path of source will be passed here
        if sql_file.endswith(".sql"):
            file1_path = os.path.join(folder1_path, sql_file)
            file2_path = os.path.join(folder2_path, sql_file)

            sp_name = sql_file

            present_in_folder1 = os.path.exists(file1_path)
            present_in_folder2 = os.path.exists(file2_path)
            content_comparison = ""
            diff_file = ""

            if present_in_folder1 and present_in_folder2:
                try:
                    with open(file1_path, "r", encoding="utf-8") as file:
                        file1_contents = file.read().upper()
                except UnicodeError:
                    with open(file1_path, "r", encoding="utf-16") as file:
                        file1_contents = file.read().upper()
                # normalized_sql_1 = normalize_sql(file1_contents)
                # file1_nocomments = strip_sql_comments(normalized_sql_1)

                try:
                    with open(file2_path, "r", encoding="utf-8") as file:
                        file2_contents = file.read().upper()
                except UnicodeError:
                    with open(file2_path, "r", encoding="utf-16") as file:
                        file2_contents = file.read().upper()

                # normalized_sql_2 = normalize_sql(file2_contents)
                # file2_nocomments = strip_sql_comments(normalized_sql_2)

                if difference(
                    file1_path, file2_path
                ):
                    content_comparison = "Equal"
                else:
                    content_comparison = "Different"
                    diff_html = generate_html_diff(
                        file1_contents, file2_contents, folder1_path, folder2_path
                    )
                    diff_filename = os.path.join(output_dir, f"diff_{sql_file}.html")

                    with open(diff_filename, "w") as diff_file:
                        diff_file.write(diff_html)
                    diff_file = diff_filename
            else:
                content_comparison = "Missing in one of the folders"
                if not present_in_folder1:
                    diff_file = f"Missing in {os.path.basename(folder1_path)}"
                else:
                    diff_file = f"Missing in {os.path.basename(folder2_path)}"

            comparison_results.append(
                [
                    sp_name,
                    present_in_folder1,
                    present_in_folder2,
                    content_comparison,
                    diff_file,
                ]
            )

    # DONE: add functionality where it shows what sp is not present in which db
    for sql_file in os.listdir(folder2_path):
        sp_name = sql_file
        if sql_file.endswith(".sql"):
            if (
                os.path.exists(os.path.join(folder2_path, sql_file)) is True
                and os.path.exists(os.path.join(folder1_path, sql_file)) is not True
            ):
                print(f"{sql_file} present in target but not in source")
                comparison_results.append(
                    [
                        sp_name,
                        os.path.exists(os.path.join(folder1_path, sql_file)),
                        os.path.exists(os.path.join(folder2_path, sql_file)),
                        f"Missing in {os.path.basename(folder1_path)}",
                    ]
                )

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        "SP Name",
        f"Present in {os.path.basename(folder1_path)}",
        f"Present in {os.path.basename(folder2_path)}",
        "Content Comparison",
        "Diff File",
    ]

    ws.append(headers)

    for row in comparison_results:
        ws.append(row)

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=len(comparison_results) + 1), start=2
    ):
        if row[3].value == "Different":
            row[3].fill = different_color
        if row[1].value is False:
            row[1].fill = missing_color
        if row[2].value is False:
            row[2].fill = missing_color

    num_files_absent = sum(not row[2] for row in comparison_results)
    num_files_different = sum(row[3] == "Different" for row in comparison_results)
    total_files_scanned = len(comparison_results)

    wb.save(excel_output_file)
    print(YELLOW + "\n\nSummary: " + RESET)
    print(
        "Number of files"
        + RED
        + " Absent "
        + RESET
        + f"in Target Database: {num_files_absent}"
    )
    print(
        "Number of files with"
        + RED
        + " Unequal Content "
        + RESET
        + f"in Target Database: {num_files_different}"
    )
    print(f"Total files scanned: {total_files_scanned}\n\n")
    print(f"Diff Files are stored in {os.path.abspath(output_dir)}")
    print(
        f"Excel file generated and stored in {os.path.dirname(os.path.abspath(excel_output_file))}\n\n"
    )
    # Open the folder insted of the excel file
    os.system(f'explorer /select, "{os.path.abspath(excel_output_file)}"')

    # END OF APP 3 ###################################################################


def get_unique_list(non_unique_list):
    unique_list = []

    for number in non_unique_list:
        if number in unique_list:
            continue
        else:
            unique_list.append(number)
    return unique_list


def app4():
    try:
        num_target_dbs = input(
            "Enter" + GREEN + " Number of Databases " + RESET + "you want to compare:\t"
        )

        if not num_target_dbs:
            print(
                RED
                + "\n\nError: "
                + RESET
                + "No input provided. Exiting the program...\n\n"
            )
            sys.exit(1)

        try:
            num_target_dbs = int(num_target_dbs)
        except ValueError:
            print(
                RED
                + "\n\nError: "
                + RESET
                + "Invalid input. Please enter a valid positive integer.\n\n"
            )
            sys.exit(1)

        target_db_dirs = []

        for i in range(num_target_dbs):
            while True:
                target_db_dir = input(
                    "\nEnter "
                    + GREEN
                    + "database directory location "
                    + RESET
                    + f"for target database number {i+1}:\t"
                )

                if not target_db_dir.strip():
                    print(
                        RED
                        + "Error: "
                        + RESET
                        + "No input provided. Please enter a directory location.\n"
                    )
                elif not os.path.exists(target_db_dir):
                    print(
                        RED
                        + "Error: "
                        + RESET
                        + f"The directory '{target_db_dir}' does not exist or is invalid. Please enter a valid directory location.\n"
                    )
                else:
                    target_db_dirs.append(target_db_dir)
                    break  # Valid input, exit the loop

        sp_data = []

        # Get the list of SQL files from all target database directories
        unique_file_list = set()
        for target_db_dir in target_db_dirs:
            sql_files = [f for f in os.listdir(target_db_dir) if f.endswith(".sql")]
            unique_file_list.update(sql_files)

        print(f"Total unique files found: {len(unique_file_list)}")

        for sql_file in unique_file_list:
            sp_name = os.path.splitext(sql_file)[0]  # Remove the .sql extension
            sp_info = {"SP Name": sp_name}

            for target_db_dir in target_db_dirs:
                target_sql_path = os.path.join(target_db_dir, sql_file)
                if os.path.exists(target_sql_path):
                    sp_info[os.path.basename(target_db_dir)] = "PRESENT"
                else:
                    sp_info[os.path.basename(target_db_dir)] = "ABSENT"

            sp_data.append(sp_info)

        df = pd.DataFrame(sp_data)

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
        output_excel_path = f"SP_Comparison_Report_{timestamp}.xlsx"
        df.to_excel(output_excel_path, index=False)

        wb = load_workbook(output_excel_path)
        ws = wb.active

        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
        ):
            for cell in row:
                if cell.value == "ABSENT":
                    cell.fill = PatternFill(
                        start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
                    )

        wb.save(output_excel_path)
        excel_absolute_path = os.path.abspath(output_excel_path)
        print(
            GREEN
            + "\nSuccess: "
            + RESET
            + f"Excel Report has been generated at {excel_absolute_path}\n"
        )
        os.system(f'explorer /select, "{os.path.abspath(output_excel_path)}"')
        input("Press Enter to exit...")

    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")


# #########################


def app5():
    try:
        server = input("Enter Server Address:\t")
        database = input("Enter Database:\t")
        username = input("Enter Username:\t")
        password = input("Enter Password:\t")
        
        sps: list[str] = fetch_stored_procedures(server, database, username, password)
        
        total_files = len(sps)

        ignore_file_path = input("Drag and drop the ignore file:\t")
        filelist_after_excl = ignore(ignore_file_path, sps)

        num_files_aft_excl = len(filelist_after_excl)

        num_excl_files = (total_files - num_files_aft_excl)
        
        # excluded_files = sps - filelist_after_excl

        print("\n\nSummary:\n")
        print(f"Number of files before exclusion: {len(sps)}")
        print(f"File list before exclusion: {sps}")
        print("\n\n")

        print(f"Number of files after exclusion: {num_files_aft_excl}")
        print(f"File list after exclusion: {filelist_after_excl}\n\n")
        print(f"Number of files excluded: {num_excl_files}")
        # print(f"Excluded files: {excluded_files}")
        

        wb = openpyxl.Workbook()
        sheet = wb.active

        # Headers
        sheet["A1"] = "Files in Database"
        sheet["B1"] = "Included or Excluded"

        # Data
        row = 2
        for sp in sps:
            if sp in filelist_after_excl:
                sheet.append([sp, "Included"])
            else:
                sheet.append([sp, "Excluded"])
            row += 1
        
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
        ex_file_name = f"Backup_File_Statistics_{timestamp}.xlsx"

        wb.save(ex_file_name)
        
        os.system(f'explorer /select,"{os.path.abspath(ex_file_name)}"')

    except Exception as e:
        print(f"{str(e)}")


# END OF APP 5 ###############


def app6():
    print("App 6")

    folder = input("Enter folder location:\t")

    file_list = os.listdir(folder)

    print("\nType 'done' and press enter when done entering patterns\n")
    n = 1
    patterns = []

    while n > 0:
        pt = input("Enter pattern:\t")
        if pt != "done":
            patterns.append(pt)
        else:
            break
        pass
    ignore_patterns_list = [r".*_" + item.upper() + ".*" for item in patterns]
    ignore_pattern = "|".join(ignore_patterns_list)
    ignore_pattern = f"({ignore_pattern})"
    excluded_files = []
    included_files = []
    for file in file_list:
        if re.match(ignore_pattern, file):
            excluded_files.append(file)
        else:
            included_files.append(file)

    print(f"Files without your pattern:\n {included_files}")

    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Included Files"
    ws["B1"] = "Excluded Files (pattern matched)"
    column_letter1 = "A"
    for file in included_files:
        cell = ws[column_letter1 + str(ws.max_row + 1)]
        cell.value = file

    column_letter2 = "B"
    for file in excluded_files:
        cell = ws[column_letter2 + str(ws.max_row + 1)]
        cell.value = file

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
    excel_file_name = f"Inluded_Excluded_Files_{timestamp}.xlsx"

    wb.save(excel_file_name)


# ###############


def fetch_sp_content(sp_name, server, database, username, password) -> str:
    connection = connect_to_server(server, database, username, password)
    cursor = connection.cursor()

    # query = f"sp_helptext {sp_name}"
    query = f"sp_helptext [{sp_name}]"

    cursor.execute(query)

    sp_content_list = cursor.fetchall()
    sp_content = "\n".join(line[0].strip() for line in sp_content_list)

    return sp_content


def app2_1() -> None:
    print("SP Analyzer Online")

    print("Enter Source details:\t")
    server = input("Enter Server Address:\t")
    database = input("Enter Database Name:\t")
    username = input("Enter User Name:\t")
    password = input("Enter Password:\t")

    print("Enter Target Database Details:\t")
    print(
        "The program will continue to ask for information until you type 'done' and press Enter when it asks.\n\n"
    )

    stop_key = "not done"
    count = 1

    target_db_details = {}

    while stop_key != "done":
        try:
            target_db_details[count] = {}
            target_db_details[count]["server"] = input(
                "Enter Target Database Server Address:\t"
            )
            target_db_details[count]["database"] = input(
                "Enter Target Database Database Name:\t"
            )
            target_db_details[count]["username"] = input(
                "Enter Target Database Username:\t"
            )
            target_db_details[count]["password"] = input(
                "Enter Target Database Password:\t"
            )

            stop_key = input(
                "Press enter to continue to the next target database. Type 'done' and press Enter if you're done.\t"
            )

            count = count + 1
        except Exception as e:
            print("Error: " + str(e))
    # While loop ends
    source_sps = fetch_stored_procedures(server, database, username, password)
    total_files = len(source_sps)

    ignore_file_path = input("Drag and drop the ignore file:\t")
    filt_source_sps = ignore(ignore_file_path, source_sps)

    num_files_aft_excl = len(filt_source_sps)

    num_excl_files = (total_files - num_files_aft_excl)

    sp_data = []

    # Loop through source SPs
    for sp in tqdm(filt_source_sps):
        sp_info = {"SP Name": sp}

        # Loop through target DBs
        if target_db_details.values() is not None:
            for target_db in target_db_details.values():
                source_sp_content = fetch_sp_content(
                    sp, server, database, username, password
                )
                try:
                    target_sp_content = fetch_sp_content(
                        sp,
                        target_db["server"],
                        target_db["database"],
                        target_db["username"],
                        target_db["password"],
                    )

                    stripped_source_sp = strip_sql_comments(source_sp_content).upper()
                    stripped_target_sp = strip_sql_comments(target_sp_content).upper()

                    if stripped_source_sp == stripped_target_sp:
                        sp_info[target_db["database"]] = "PRESENT & EQUAL"
                    else:
                        sp_info[target_db["database"]] = "PRESENT & UNEQUAL"

                except Exception:
                    sp_info[target_db["database"]] = "ABSENT"
            else:
                print("Target DB Details are None")

        sp_data.append(sp_info)

    df = pd.DataFrame(sp_data)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
    output_excel_file = f"SP_Comparison_Report_Online_{timestamp}.xlsx"
    df.to_excel(output_excel_file, index=False)
    wb = load_workbook(output_excel_file)
    ws = wb.active

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
    ):
        for cell in row:
            if cell.value == "PRESENT & UNEQUAL":
                cell.fill = PatternFill(
                    start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"
                )
            elif cell.value == "ABSENT":
                cell.fill = PatternFill(
                    start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
                )

    # Save the modified workbook
    wb.save(output_excel_file)
    print(
        f"\n\nExcel file successfully created: {os.path.abspath(output_excel_file)}\n\n"
    )
    print(f"Total Files: {total_files}")
    print(f"Files Excluded: {num_excl_files}")
    print(f"Files Considered: {num_files_aft_excl}")

    os.system(f'explorer /select,"{os.path.abspath(output_excel_file)}"')
# ################


def get_database_details():
    database_details = []
    while True:
        server = input("Enter Database Server Address (or 'done' to finish):\t")
        if server.lower() == 'done':
            break
        database = input("Enter Database Name:\t")
        username = input("Enter Database Username:\t")
        password = input("Enter Database Password:\t")
        database_details.append({
            "server": server,
            "database": database,
            "username": username,
            "password": password
        })
    return database_details


def app2_2() -> None:
    print("\n\nOnline SP Presence Analyzer for Multiple Databases\n\n")
    print("This program will run as long as you don't enter 'done' when asked for")

    database_details = get_database_details()
    total_files = 0

    all_sp_names = []

    for database_detail in database_details:
        server = database_detail['server']
        database = database_detail['database']
        username = database_detail['username']
        password = database_detail['password']

        sp_names = fetch_stored_procedures(server, database, username, password)

        all_sp_names.extend(sp_names)

    # Remove duplicates by converting the list to a set and back to a list
    superset_sp_names = list(set(all_sp_names))  # unique but not filtered
    total_files = len(superset_sp_names)

    ignore_file_path = input("Enter complete path of ignore file or drag and drop the ignore file:\t")
    filt_sp_names = ignore(ignore_file_path, superset_sp_names)

    num_files_aft_excl = len(filt_sp_names)
    num_excl_files = total_files - num_files_aft_excl

    sp_data = []

    for sp in tqdm(filt_sp_names):
        sp_info = {"SP Name": sp}
        for database_detail in database_details:
            server = database_detail['server']
            database = database_detail['database']
            username = database_detail['username']
            password = database_detail['password']

            current_db_sp_list = fetch_stored_procedures(server, database, username, password)
            if sp in current_db_sp_list:
                sp_info[database_detail["database"]] = "PRESENT"
            else:
                sp_info[database_detail["database"]] = "ABSENT"

        sp_data.append(sp_info)

    df = pd.DataFrame(sp_data)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
    output_excel_file = f"SP_Presence_Report_Online_{timestamp}.xlsx"
    df.to_excel(output_excel_file, index=False)
    wb = load_workbook(output_excel_file)
    ws = wb.active

    # Apply cell coloring based on the cell values
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
    ):
        for cell in row:
            if cell.value == "ABSENT":
                cell.fill = PatternFill(
                    start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
                )
    # Save the modified workbook
    wb.save(output_excel_file)
    print(
        f"\n\nExcel file successfully created: {os.path.abspath(output_excel_file)}\n\n"
    )
    print(f"Total Files: {total_files}")
    print(f"Files Excluded: {num_excl_files}")
    print(f"Files Considered: {num_files_aft_excl}")

    os.system(f'explorer /select,"{os.path.abspath(output_excel_file)}"')


# ###################
def app2_3() -> None:
    print("\n\nOffline SP Presence Analyzer for Multiple Databases\n\n")
    print("This program will run as long as you don't enter 'done' when asked for")

    truth: bool = True
    db_list: list[str] = []

    while truth:
        temporary: str = input("\nEnter directory location for next database:\t")
        if temporary != 'done':
            db_list.append(temporary)
        else:
            truth = False
    # while loop ends

    non_unique_sp_list: list[str] = []
    for db in db_list:
        for sp in os.listdir(db):
            non_unique_sp_list.append(sp)
    # for loop ends

    unique_sp_list: list[str] = list(set(non_unique_sp_list))
    total_files_before_exclusion = len(unique_sp_list)

    ignore_file_path = input("Enter complete path of ignore file or drag and drop the ignore file:\t")

    with open(ignore_file_path, "r") as f:
        ignore = f.read().splitlines()

    additional_ignore_list = ['*.mig*', '*sqlquery*', '*2022*', '*2023*']

    ignore.extend(additional_ignore_list)
    ignore_patterns_list = [r".*_" + item.upper() + ".*" for item in ignore]


    # Create a regular expression pattern to match ignore patterns
    ignore_pattern = "|".join(ignore_patterns_list)
    ignore_pattern = f"({ignore_pattern})"

    # Exclusion based on pattern matching file
    unique_sp_list_excl = [sp for sp in unique_sp_list if not re.match(ignore_pattern, sp.upper())]

    total_files_after_exclusion = len(unique_sp_list_excl)
    number_of_files_excluded = total_files_before_exclusion - total_files_after_exclusion



    data: dict[str, list[str]] = {'SP Name': unique_sp_list_excl}
    for db in db_list:
        sps_in_current_db: list[str] = os.listdir(db)
        presence: list[bool] = [sp in sps_in_current_db for sp in unique_sp_list_excl]
        data[os.path.basename(db)] = ['PRESENT' if p else 'ABSENT' for p in presence]
    # for loop ends

    df: pd.DataFrame = pd.DataFrame(data)
    print(type(df))

    timestamp: str = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
    excel_file_name: str = f"SP_Presence_Offline_Report_{timestamp}.xlsx"

    df.to_excel(excel_file_name, index=False)
    print(f"\n\nExcel Report generated: {os.path.abspath(excel_file_name)}\n\n")


    # Load the existing workbook and sheet
    wb: Workbook = load_workbook(excel_file_name)
    ws = wb.active
    # Apply cell coloring based on the cell values
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
    ):
        for cell in row:
            if cell.value == "ABSENT":
                cell.fill = PatternFill(
                    start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
                    )

    # Save the modified workbook
    wb.save(excel_file_name)

    print("\n\nSummary:\n")
    print(f"Number of files before exclusion:\t{total_files_before_exclusion}")
    print(f"Number of files after exclusion:\t{total_files_after_exclusion}")
    print(f"Number of excluded files:\t{number_of_files_excluded}")





# ###################



def app2():
    online = input("\n\n1. Online SP Comparator among multiple databases. Compares contents as well. Comparison done against a source.\n2. Online SP Presence Analyser among multiple databases. Makes a superset of all SPs and checks which SP is absent in what database.\n3. Offline SP Presence Analyzer for multiple databases. Superset is made. No content comparison is done.\n4. Offline SP Comparator among multiple databases. Comparison done against a source.\nEnter your choice:\t").strip()
    online = int(online)

    match online:
        case 1:
            app2_1()
        case 2:
            app2_2()
        case 3:
            app2_3()
        case 4:
            app2_4()
        case _:
            print("Invalid choice. Exiting...")
            sys.exit(1)


# ############


def extract_schema_from_sql(sql_file):
    # Read the SQL file and extract schema information
    schema = {}
    with open(sql_file, 'r') as file:
        sql_content = file.read()

    # Use regular expressions to extract schema details
    # Modify the regex patterns as needed to match your SQL file format
    table_pattern = r'CREATE TABLE \[dbo\]\.\[([^\]]+)\]'
    column_pattern = r'\[([^\]]+)\] ([^\s]+)'

    tables = re.findall(table_pattern, sql_content)

    for table in tables:
        schema[table] = []
        column_matches = re.findall(column_pattern, sql_content)
        for match in column_matches:
            column_name, column_type = match
            schema[table].append((column_name, column_type))

    return schema

def compare_schemas(source_folder, target_folders):
    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Schema Comparison Summary"
    summary_sheet.append(["File Name", "Database Name", "Table Name", "Status"])

    # Compare schemas for each target folder
    for target_folder in target_folders:
        target_sql_file = os.path.join(target_folder, f"{target_folder}.sql")
        if not os.path.exists(target_sql_file):
            continue

        target_schema = extract_schema_from_sql(target_sql_file)

        # Iterate through SQL files in the source folder
        for root, _, files in os.walk(source_folder):
            for file in files:
                if file.endswith(".sql"):
                    source_sql_file = os.path.join(root, file)
                    source_schema = extract_schema_from_sql(source_sql_file)
                    db_name = os.path.basename(target_folder)
                    file_name = os.path.basename(source_sql_file)

                    missing_tables, different_tables = compare_schemas(source_schema, target_schema)

                    # Add comparison results to the summary sheet
                    for table in missing_tables:
                        summary_sheet.append([file_name, db_name, table, "Missing"])

                    for table in different_tables:
                        summary_sheet.append([file_name, db_name, table, "Different"])

    # Save the Excel report
    report_file = "Schema_Comparison_Report.xlsx"
    workbook.save(report_file)

def app1_2():
    source_folder = "RFCTEST"  # Change this to the folder containing RFCTEST SQL files
    target_folders = ["CPUH", "CPUK", "CRESCENT", "DAIICT", "DCH", "HITS", "IPER", "IPR", "JECRC", "MAHER", "MIT", "PRMCEM", "PRMITR", "RCPIPER", "RCPIT"]

    compare_schemas(source_folder, target_folders)

    print("Schema comparison report saved")

# ############


def app7() -> None:
    folder = input("Enter path for folder with excel files:\t")
    
    datatype1 = ["bit", "datetime", "date", "int", "bigint", "tinyint", "time"]
    datatype2 = ["numeric", "decimal", "float"]
    datatype3 = ["varchar", "nvarchar", "char", "nchar"]
    
    # List directories
    xl_files = os.listdir(folder)
    
    for xl_file in xl_files:
        workbook = openpyxl.load_workbook(os.path.join(folder, xl_file))
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet['Q1'] = "Query"
            
            for row in range(2, sheet.max_row + 1):
                operation = str(sheet.cell(row=row, column=4).value)
                data_type = str(sheet.cell(row=row, column=8).value)
                max_length = sheet.cell(row=row, column=9).value
                # column_name = sheet.cell(row=row, column=3).value
                
                # Check the operation and data type and build SQL accordingly
                if operation == "Missing Column":
                    # if any("abc" in s for s in xs):
                    if any(data_type in s for s in datatype1):
                        sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ADD ", "[", C{row}, "]", " " , H{row})')
                    elif any(data_type in s for s in datatype2):
                        sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ADD ", "[", C{row}, "]", " ", H{row}, " (", J{row}, ",", K{row}, ")")')

                        # sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ADD ", C{row}, " " , H{row}, " (", J{row},",","K{row},")")')
                    elif any(data_type in s for s in datatype3):
                        
                        if max_length == -1:
                            sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ADD ", "[", C{row}, "]", " " , H{row}," (max)")')
                        else:
                            sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ADD ", "[", C{row}, "]", " " , H{row}, " (",I{row},")")')
                    else:
                        print("Unknown data type")
                elif operation == "Different Specification":
                    if any(data_type in s for s in datatype1):
                        sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ALTER COLUMN ", "[", C{row}, "]", " " , H{row})')
                    elif any(data_type in s for s in datatype2):
                            sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ALTER COLUMN ", "[", C{row}, "]", " ", H{row}, " (", J{row}, ",", K{row}, ")")')
                    elif any(data_type in s for s in datatype3):
                        
                        if max_length == -1:
                            #                                           =CONCAT("ALTER TABLE ",B12," ALTER COLUMN ",C12," ",H12,"(max)")
                            sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ALTER COLUMN ", "[", C{row}, "]", " " , H{row}, " (max)")')
                        else:
                            sheet.cell(row=row, column=17, value=f'=CONCATENATE("ALTER TABLE ", B{row}, " ALTER COLUMN ", "[", C{row}, "]", " " , H{row}, " (", I{row}, ")")')
                elif operation == "Missing Table":
                    print("Table is missing. Skipping operation.")
                else:
                    print(f"Type of error unknown for row {row} in sheet {sheet_name}")
                    
        workbook.save(os.path.join(folder, xl_file))





# column D: Type of Error
# column H: source data type
# query in column Q
# inside each sheet first check type of error

# if type of error is missing columns
# then source data type
# if data type is bit,datetime,date,int
# then alter table table_name add column_name datatype
# if data type is numeric,decimal
# then write query: alter table table_name add column_name datatype(Source Numeric Precision,Source Numeric Scale)
# if data type nvarchar or varchar
# alter table table_name add column_name datatype(Source Max Length) --if Source Max Length = -1 max

# else if type of error is diff specification
# then check data type
# if data type is bit,datetime,date,int
# alter table table_name alter column column_name datatype
# if data type is numeric,decimal
# alter table table_name alter column column_name datatype(Source Numeric Precision,Source Numeric Scale)
# if data type nvarchar or varchar
# alter table table_name alter column column_name datatype(Source Max Length) --if Source Max Length = -1 max

# else if type of error is missing table
# then do nothing

# below is for different specification



# ##########################

def main():
    try:
        print("\n\nPlease enter what program you wish to execute: \n\n")
        print(
            YELLOW
            + "1. Database Schema Analyzer: "
            + RESET
            + "Compare database schemas from multiple sources and generate Excel reports\n"
        )
        print(
            YELLOW
            + "2. Stored Procedure Analyzer: "
            + RESET
            + "Compare Stored Procedures from multiple databases on your system and generate Excel report\n"
        )
        print(
            YELLOW
            + "3. Stored Procedure Comparator + Diff Generator: "
            + RESET
            + "Compare Stored Procedures between two databases on your system, generate Excel reports, and store differential files in HTML format for visualizing differences\n"
        )
        print(
            YELLOW
            + "4. Stored Procedure Cross-Database Analyzer: "
            + RESET
            + "Examine the presence of stored procedures across multiple databases, ensuring their mutual existence.\n"
        )
        print(YELLOW + "\n5. Excluded File Statistics: " + RESET)
        print(YELLOW + "\n6. Exclude Files Based on Pattern Matching: " + RESET)
        print(YELLOW + "\n7. Generate Alter Query" + RESET)

        choice = int(input("Enter your choice:\t"))
        print(f"You have selected option: {choice}.\n")

        if choice == 1:
            app1()
        elif choice == 2:
            app2()
        elif choice == 3:
            app3()
        elif choice == 4:
            app4()
        elif choice == 5:
            app5()
        elif choice == 6:
            app6()
        elif choice == 7:
            app7()
        elif choice == 8:
            app1_2()
        else:
            print(
                RED + "Error: " + RESET + "Please select a valid choice from 1 to 6\n"
            )
    except ValueError:
        print(RED + "\nError: " + RESET + "Please enter a valid numeric choice.\n")
    except Exception as e:
        print(f"Error occurred: {str(e)}")


if __name__ == "__main__":
    main()
