import openpyxl
from openpyxl import Workbook

excel_file = "C:\\Users\\swarn\\github\\Sequel\\SP_Comparison_Report_2023-09-04_18.35.51.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

absent_data = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row) if ws.cell(i,3).value == 'ABSENT']
unequal_data =[ws.cell(row=i, column=3).value for i in range(3, ws.max_row) if ws.cell(i,3).value == 'PRESENT & UNEQUAL']

print(len(absent_data))
print(len(unequal_data))

row_name = [ws.cell(1,i).value for i in range(2, ws.max_column+1)]
print(row_name)

