#Push
import os

from PIL import Image
import customtkinter
import openpyxl
from openpyxl.styles import PatternFill
import pyodbc



class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("image_example.py")
        self.geometry("700x450")
        self.after(0, lambda: self.wm_state("zoomed"))

        # Font for labels and entries
        # label_font = CTkFont(family="Helvetica", size=12, weight="bold")
        # entry_font = CTkFont(family="Helvetica", size=12)

        self.font = customtkinter.CTkFont(family="JetBrains Mono", size=18)
        font = customtkinter.CTkFont(family="JetBrains Mono", size=18)

        # set grid layout 1x2
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # load images with light and dark mode image
        image_path = "C:\\Users\\swarn\\Downloads\\examples\\test_images"
        self.logo_image = customtkinter.CTkImage(
            Image.open(os.path.join(image_path, "CustomTkinter_logo_single.png")),
            size=(26, 26),
        )
        self.large_test_image = customtkinter.CTkImage(
            Image.open(os.path.join(image_path, "large_test_image.png")),
            size=(500, 150),
        )
        self.image_icon_image = customtkinter.CTkImage(
            Image.open(os.path.join(image_path, "image_icon_light.png")), size=(20, 20)
        )
        self.home_image = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "home_dark.png")),
            dark_image=Image.open(os.path.join(image_path, "home_light.png")),
            size=(20, 20),
        )
        self.chat_image = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "chat_dark.png")),
            dark_image=Image.open(os.path.join(image_path, "chat_light.png")),
            size=(20, 20),
        )
        self.add_user_image = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "add_user_dark.png")),
            dark_image=Image.open(os.path.join(image_path, "add_user_light.png")),
            size=(20, 20),
        )

        # create navigation frame
        self.navigation_frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(
            self.navigation_frame,
            text="  Image Example",
            image=self.logo_image,
            compound="left",
            font=customtkinter.CTkFont(size=15, weight="bold"),
        )
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)

        self.home_button = customtkinter.CTkButton(
           class ScrollableLabelButtonFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master, command=None, **kwargs):
        super().__init__(master, **kwargs)
        self.grid_columnconfigure(0, weight=1)

        self.command = command
        self.radiobutton_variable = customtkinter.StringVar()
        self.label_list = []
        self.button_list = []

    def add_item(self, item, image=None):
        label = customtkinter.CTkLabel(self, text=item, image=image, compound="left", padx=5, anchor="w")
        button = customtkinter.CTkButton(self, text="Command", width=100, height=24)
        if self.command is not None:
            button.configure(command=lambda: self.command(item))
        label.grid(row=len(self.label_list), column=0, pady=(0, 10), sticky="w")
        button.grid(row=len(self.button_list), column=1, pady=(0, 10), padx=5)
        self.label_list.append(label)
        self.button_list.append(button)

    def remove_item(self, item):
        for label, button in zip(self.label_list, self.button_list):
            if item == label.cget("text"):
                label.destroy()
                button.destroy()
                self.label_list.remove(label)
                self.button_list.remove(button)
                return self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="Home",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            image=self.home_image,
            anchor="w",
            command=self.home_button_event,
        )
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = customtkinter.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="Frame 2",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            image=self.chat_image,
            anchor="w",
            command=self.frame_2_button_event,
        )
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.frame_3_button = customtkinter.CTkButton(
            self.navigation_frame,
            corner_radius=0,
            height=40,
            border_spacing=10,
            text="Frame 3",
            fg_color="transparent",
            text_color=("gray10", "gray90"),
            hover_color=("gray70", "gray30"),
            image=self.add_user_image,
            anchor="w",
            command=self.frame_3_button_event,
        )
        self.frame_3_button.grid(row=3, column=0, sticky="ew")

        self.appearance_mode_menu = customtkinter.CTkOptionMenu(
            self.navigation_frame,
            values=["Light", "Dark", "System"],
            command=self.change_appearance_mode_event,
        )
        self.appearance_mode_menu.grid(row=6, column=0, padx=20, pady=20, sticky="s")
        ## ------------------------------------
        # create home frame
        self.home_frame = customtkinter.CTkFrame(
            self, corner_radius=0, fg_color="transparent"
        )
        self.home_frame.grid_columnconfigure(0, weight=1)

        self.home_frame_large_image_label = customtkinter.CTkLabel(
            self.home_frame, text="", image=self.large_test_image
        )
        self.home_frame_large_image_label.grid(row=0, column=0, padx=20, pady=10)

        self.home_frame_button_1 = customtkinter.CTkButton(
            self.home_frame, text="", image=self.image_icon_image
        )
        self.home_frame_button_1.grid(row=1, column=0, padx=20, pady=10)
        self.home_frame_button_2 = customtkinter.CTkButton(
            self.home_frame,
            text="CTkButton",
            image=self.image_icon_image,
            compound="right",
        )
        self.home_frame_button_2.grid(row=2, column=0, padx=20, pady=10)
        self.home_frame_button_3 = customtkinter.CTkButton(
            self.home_frame,
            text="CTkButton",
            image=self.image_icon_image,
            compound="top",
        )
        self.home_frame_button_3.grid(row=3, column=0, padx=20, pady=10)
        self.home_frame_button_4 = customtkinter.CTkButton(
            self.home_frame,
            text="CTkButton",
            image=self.image_icon_image,
            compound="bottom",
            anchor="w",
        )
        self.home_frame_button_4.grid(row=4, column=0, padx=20, pady=10)

        # create second frame
        self.second_frame = customtkinter.CTkFrame(
            self, corner_radius=0, fg_color="transparent"
        )
        self.second_frame.grid_columnconfigure(0, weight=1)

        # Enter Server:
        self.second_frame_label_server = customtkinter.CTkLabel(
            self.second_frame, text="Server Address: ", font=font
        )
        # self.second_frame_label_server.grid(row=2, column=0, padx=20, pady=10, sticky="w")
        self.second_frame_label_server.place(relx=0.10, rely=0.10)

        self.second_frame_entry_server = customtkinter.CTkEntry(
            self.second_frame, font=font
        )
        # self.second_frame_entry_server.grid(row=2, column=1, padx=40, pady=10)
        self.second_frame_entry_server.place(relx=0.60, rely=0.10)

        # Enter Source DB
        self.second_frame_label_source = customtkinter.CTkLabel(
            self.second_frame, text="Source Database Name: ", font=font
        )
        # self.second_frame_label_source.grid(row=3, column=0, padx=20, pady=10, sticky="w")
        self.second_frame_label_source.place(relx=0.10, rely=0.20)

        self.second_frame_entry_source = customtkinter.CTkEntry(
            self.second_frame, font=font
        )
        # self.second_frame_entry_source.grid(row=3, column=1, padx=20, pady=10)
        self.second_frame_entry_source.place(relx=0.60, rely=0.20)

        ## Enter Comma Separated List of Target Databases
        self.second_frame_label_target = customtkinter.CTkLabel(
            self.second_frame,
            text="Source Target DB List (comma-separated): ",
            font=font,
        )
        # self.second_frame_label_target.grid(row=4, column=0, padx=20, pady=10, sticky="w")
        self.second_frame_label_target.place(relx=0.10, rely=0.30)

        self.second_frame_entry_target = customtkinter.CTkEntry(
            self.second_frame, font=font
        )
        # self.second_frame_entry_target.grid(row=4, column=1, padx=20, pady=10)
        self.second_frame_entry_target.place(relx=0.60, rely=0.30)

        ## Enter usernames
        self.second_frame_label_username = customtkinter.CTkLabel(
            self.second_frame, text="Username: ", font=font
        )
        # self.second_frame_label_username.grid(row=5, column=0, padx=20, pady=10, sticky="w")
        self.second_frame_label_username.place(relx=0.10, rely=0.40)

        self.second_frame_entry_username = customtkinter.CTkEntry(
            self.second_frame, font=font
        )
        # self.second_frame_entry_username.grid(row=5, column=1, padx=20, pady=10)
        self.second_frame_entry_username.place(relx=0.60, rely=0.40)

        ## Enter password
        self.second_frame_label_password = customtkinter.CTkLabel(
            self.second_frame, text="Password: ", font=font
        )
        # self.second_frame_label_password.grid(row=6, column=0, padx=20, pady=10, sticky="w")
        self.second_frame_label_password.place(relx=0.10, rely=0.50)

        self.second_frame_entry_password = customtkinter.CTkEntry(
            self.second_frame, show="*", font=font
        )
        # self.second_frame_entry_password.grid(row=6, column=1, padx=20, pady=10)
        self.second_frame_entry_password.place(relx=0.60, rely=0.50)

        ## Button
        self.second_frame_button = customtkinter.CTkButton(
            self.second_frame,
            text="Submit",
            compound="bottom",
            font=font,
            command=self.app2_submit_info,
        )
        self.second_frame_button.place(relx=0.5, rely=0.85, anchor=customtkinter.CENTER)


        # create second frame
        self.third_frame = customtkinter.CTkFrame(
            self, corner_radius=0, fg_color="transparent"
        )
        # self.third_frame.grid_columnconfigure(0, weight=1)
        # self.third_frame.grid_rowconfigure(2, weight=2)
        #

    def app2_submit_info(self):
        app2_server = self.second_frame_entry_server.get()
        app2_source_db = self.second_frame_entry_source.get()
        app2_target_dbs = self.second_frame_entry_target.get()
        app2_username = self.second_frame_entry_username.get()
        app2_password = self.second_frame_entry_password.get()

        # call
        self.app2_schema_compare(
            app2_server, app2_source_db, app2_target_dbs, app2_username, app2_password
        )

    #########
    # Function to download schema
    ###########
    # 02 - data then comes here
    def app2_fetch_schema(self, server, database, username, password):
        schema_info = {}

        # Establish connection:
        connection_string = f"DRIVER=SQL Server;SERVER={server};DATABASE={database};UID={username};PWD={password}"
        connection = pyodbc.connect(connection_string)
        cursor = connection.cursor()

        # Fetch Schema
        schema_query = """
        SELECT
            t.TABLE_SCHEMA,
            t.TABLE_NAME,
            c.COLUMN_NAME,
            c.DATA_TYPE,
            c.CHARACTER_MAXIMUM_LENGTH,
            c.NUMERIC_PRECISION,
            c.NUMERIC_SCALE
        FROM INFORMATION_SCHEMA.TABLES AS t
        JOIN INFORMATION_SCHEMA.COLUMNS AS c ON t.TABLE_SCHEMA = c.TABLE_SCHEMA AND t.TABLE_NAME = c.TABLE_NAME
        WHERE t.TABLE_TYPE = 'BASE TABLE'
        ORDER BY t.TABLE_SCHEMA, t.TABLE_NAME, c.ORDINAL_POSITION
        """

        cursor.execute(schema_query)
        rows = cursor.fetchall()

        for row in rows:
            table_schema = row.TABLE_SCHEMA
            table_name = row.TABLE_NAME
            column_name = row.COLUMN_NAME
            data_type = row.DATA_TYPE
            max_length = row.CHARACTER_MAXIMUM_LENGTH
            numeric_precision = row.NUMERIC_PRECISION
            numeric_scale = row.NUMERIC_SCALE

            schema_info.setdefault(table_schema, {}).setdefault(table_name, []).append(
                {
                    "column_name": column_name,
                    "data_type": data_type,
                    "max_length": max_length,
                    "numeric_precision": numeric_precision,
                    "numeric_scale": numeric_scale,
                }
            )

        connection.close()

        return schema_info  # Returns a dictionary with the fetched schema

    # 01 - First the data comes here
    def app2_schema_compare(self, server, source_db, target_dbs, username, password):
        # Fetch schema - data goes to the fetch schema function
        source_schema_info = self.app2_fetch_schema(
            server, source_db, username, password
        )  # source schema is fetched and stored in this variable

        # Create an Excel workbook
        workbook = openpyxl.Workbook()

        # Loop through target databases
        for target_db in target_dbs.split(","):
            # schema of target databases is fetched one by one
            target_schema_info = self.app2_fetch_schema(
                server, target_db, username, password
            )
            # one by one the data of each target is sent for comparison
            comparison_results = self.perform_schema_comparison(
                source_schema_info, target_schema_info
            )

            # Create a new sheet for the current target database
            sheet = workbook.create_sheet(title=target_db)

            # Write headers
            sheet["A1"] = "Schema"
            sheet["B1"] = "Table Name"
            sheet["C1"] = "Column Name"
            sheet["D1"] = "Data Type"
            sheet["E1"] = "Max Length"
            sheet["F1"] = "Numeric Precision"
            sheet["G1"] = "Numeric Scale"

            # Populate the sheet with comparison results
            row = 2
            for result in comparison_results:
                sheet[f"A{row}"] = result["schema"]
                sheet[f"B{row}"] = result["table_name"]
                sheet[f"C{row}"] = result["column_name"]
                sheet[f"D{row}"] = result["data_type"]
                sheet[f"E{row}"] = result["max_length"]
                sheet[f"F{row}"] = result["numeric_precision"]
                sheet[f"G{row}"] = result["numeric_scale"]

                if result["data_type"] == "Missing Column":
                    sheet[f"D{row}"].fill = PatternFill(
                        start_color="F9B9B7", end_color="F9B9B7", fill_type="solid"
                    )
                elif result["data_type"] == "Different Specification":
                    sheet[f"D{row}"].fill = PatternFill(
                        start_color="D1D5DE", end_color="D1D5DE", fill_type="solid"
                    )

                row += 1

        # Save the Excel file
        excel_filename = "schema_comparison_report.xlsx"
        workbook.remove(workbook.active)  # Remove the default sheet
        workbook.save(excel_filename)

        print(f"Schema comparison report saved to {excel_filename}")

    def perform_schema_comparison(self, source_schema, target_schema):
        comparison_results = []
        for schema in source_schema:
            for table_name in source_schema[schema]:
                source_columns = source_schema[schema][table_name]
                target_columns = target_schema[schema].get(table_name, [])

                for col_info_source in source_columns:
                    col_name_source = col_info_source["column_name"]
                    col_info_target = next(
                        (
                            col
                            for col in target_columns
                            if col["column_name"] == col_name_source
                        ),
                        None,
                    )

                    if col_info_target is None:
                        comparison_result = {
                            "schema": schema,
                            "table_name": table_name,
                            "column_name": col_name_source,
                            "data_type": "Missing Column",
                            "max_length": str(col_info_source),
                            "numeric_precision": "",
                            "numeric_scale": "",
                        }
                        comparison_results.append(comparison_result)

                    elif col_info_source != col_info_target:
                        comparison_result = {
                            "schema": schema,
                            "table_name": table_name,
                            "column_name": col_name_source,
                            "data_type": "Different Specification",
                            "max_length": str(col_info_source),
                            "numeric_precision": str(col_info_target),
                            "numeric_scale": "",
                        }
                        comparison_results.append(comparison_result)

        return comparison_results

    def select_frame_by_name(self, name):
        # set button color for selected button
        self.home_button.configure(
            fg_color=("gray75", "gray25") if name == "home" else "transparent"
        )
        self.frame_2_button.configure(
            fg_color=("gray75", "gray25") if name == "frame_2" else "transparent"
        )
        self.frame_3_button.configure(
            fg_color=("gray75", "gray25") if name == "frame_3" else "transparent"
        )

        # show selected frame
        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "frame_2":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.second_frame.grid_forget()
        if name == "frame_3":
            self.third_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.third_frame.grid_forget()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def frame_2_button_event(self):
        self.select_frame_by_name("frame_2")

    def frame_3_button_event(self):
        self.select_frame_by_name("frame_3")

    def change_appearance_mode_event(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)


if __name__ == "__main__":
    app = App()
    app.mainloop()
